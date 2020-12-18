## PREREQUISITES
#1.Check Chrome Version And Download Appropriate Chromedriver(HAS to be in SAME PATH).
#(Eg My Chromedriver Version Is 87.0.4280.88  And My OS is Windows So I Downloaded The Chromedriver Version 87.)

#2.Join Multiple Facebook Groups According To Your Niche And Store Their Name And Group Link In Excel File.
#(Eg My Excel File Name Is 'grouplist.xlsx'(line 28) )

#use function marketme(sheetno,startrow,endrow): to start posting your links and other text in various facebook groups.
#Eg marketme(0,1,15)..will post to groups in rows 1 to 15 of sheet 0 in our excel file.

#Developer CONTACT(JOHNTY GOMES)
#https://api.whatsapp.com/send?phone=+919773211427

import time
import openpyxl
import xlrd
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
driver=webdriver.Chrome()
driver.get("https://www.facebook.com/")
time.sleep(4)
eml=driver.find_element_by_xpath("//input[@name='email']")
eml.send_keys("YOUR_FACEBOOK_EMAIL")
pswd=driver.find_element_by_xpath("//input[@name='pass']")
pswd.send_keys("YOUR_FACEBOOK_PASSWORD")
login=driver.find_element_by_xpath("//button[@name='login']")
login.click()
time.sleep(5)

def marketme(sheetno,startrow,endrow):
    loc = 'grouplist.xlsx'
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(sheetno)

    file = loc
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[sheetno]

    for i in range(startrow,endrow):
        try:
            driver.get(sheet.cell_value(i,1))
            time.sleep(7)
            post=driver.find_element_by_xpath("//span[contains(text(),' on your mind, Guitarical?')]")
            post.click()
            time.sleep(10)
            temp1="Hello Members"+"\u2764 Of "+sheet.cell_value(i,0)+" ."
            temp=temp1+"\nhttps://www.youtube.com/watch?v=j7AGHTyP0rU\nPink Panther Guitar\u2763 Fingerstyle\u266B \u270C Tutorial\u270D And Cover For Beginners With Diagrams.\nLearn\u270D This Fingerstyle\u266B in \u0031\uFE0F\u20E3\u0030\uFE0F\u20E3 Minutes\u23F0.\nImpress Your Friends\u2764.\nGuaranteed\u0023\uFE0F\u20E3\u0031\uFE0F\u20E3\u0030\uFE0F\u20E3\u0030\uFE0F\u20E3% Results.\nRequests For Guitar/Ukulele Fingerstyle Tutorials Bollywood/Western Accepted in Hindi\u2643 as Well As English\u264F.\nWhatsapp\u2706 Call\u260E\u27A1+919773211427.\nNow Is The Time To Learn\u270D And Upgrade\u2934.\nPick Your Guitar And Get Right On To It Quick\u23F3 Before It Gets Too Late\u2620."
            post=driver.find_elements_by_xpath("//div[@contenteditable='true' and @ role='textbox']")
            post[len(post)-1].send_keys(temp)
            time.sleep(8)
            #driver.find_element_by_xpath("//div[contains(text(),'OK') and @role='button']").click()
            #time.sleep(3)
            #c=driver.find_elements_by_xpath("//input[@accept='image/*,image/heif,image/heic,video/*,video/mp4,video/x-m4v,video/x-matroska,.mkv']")
            #c[0].send_keys("D:/JOHNTY/uddgayefinaldone.mp4")
            post_btn=driver.find_element_by_xpath("//div[@aria-label='Post']")
            post_btn.click()
            print("done")
            ws.cell(row=i+1, column=3, value="SENT")
            wb.save(file)
            time.sleep(14)
        except Exception as e:
            print(e)
            terrmsg="coud nt send to "+sheet.cell_value(i,0)+"####"+str(i)
            print(terrmsg)
            #ws.cell(row=i+1, column=1, value=sheet.cell_value(i,0))
            #ws.cell(row=i+1, column=2, value=sheet.cell_value(i,1))
            #wb.save(file)  

'''
temp=temp1+"\nlink\nPapa Kehte Hain Bada Naam Karega\u2763 Guitar\u266B Fingerstyle\u270C Tutorial\u270D And Cover For Beginners.\nLearn\u270D This Song\u266B in \u0031\uFE0F\u20E3\u0033\uFE0F\u20E3 Minutes\u23F0.\nImpress Your Friends\u2764.\nGuaranteed\u0023\uFE0F\u20E3\u0031\uFE0F\u20E3\u0030\uFE0F\u20E3\u0030\uFE0F\u20E3% Results.\nRequests For Guitar/Ukulele Fingerstyle Tutorials Bollywood/Western Accepted in Hindi\u2643 as Well As English\u264F.\nWhatsapp\u2706 Call\u260E\u27A1+919773211427.\nNow Is The Time To Learn\u270D And Upgrade\u2934.\nPick Your Guitar And Get Right On To It Quick\u23F3 Before It Gets Too Late\u2620."
https://www.youtube.com/watch?v=xgJFEsLcsbI&list=PLu3gX9jLDLeL-QSBeATIhZ6zaoEGShRcU&index=2&t=0s
'''
